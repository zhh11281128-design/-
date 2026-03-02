import datetime
import os
os.environ['KIVY_GL_BACKEND'] = 'angle_sdl2'
import sys
import sqlite3
from openpyxl import Workbook, load_workbook
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.spinner import Spinner
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.core.text import LabelBase, DEFAULT_FONT
from kivy.resources import resource_add_path, resource_find
from kivy.metrics import dp
from plyer import filechooser

# ----- 解决中文显示问题（自动定位字体文件）-----
def setup_chinese_font():
    """自动检测并注册中文字体，支持打包后的资源访问"""
    possible_paths = [
        os.path.join(os.path.dirname(__file__), 'simhei.ttf'),
        os.path.join(getattr(sys, '_MEIPASS', ''), 'simhei.ttf'),
        resource_find('simhei.ttf')
    ]
    font_path = None
    for p in possible_paths:
        if p and os.path.exists(p):
            font_path = p
            break
    if not font_path:
        raise FileNotFoundError(
            "字体文件 simhei.ttf 不存在，请将其放在程序目录下，"
            "或确保打包时已包含。"
        )
    LabelBase.register(name='SimHei', fn_regular=font_path)
    LabelBase.register(name=DEFAULT_FONT, fn_regular=font_path)

setup_chinese_font()

# 全局定义年份范围（2023-2050）
START_YEAR = 2023
END_YEAR = 2050
YEAR_LIST = [str(y) for y in range(START_YEAR, END_YEAR + 1)]

# 数据库表名和列定义
TABLE_NAME = 'records'
COLUMNS = ['date', 'type', 'amount', 'event', 'party', 'platform', 'other']
CREATE_TABLE_SQL = f'''
    CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL,
        type TEXT NOT NULL,
        amount REAL NOT NULL,
        event TEXT,
        party TEXT,
        platform TEXT,
        other TEXT
    )
'''
CREATE_INDEX_SQL = f'CREATE INDEX IF NOT EXISTS idx_date ON {TABLE_NAME}(date);'


# ---------- 数据库初始化辅助函数 ----------
def init_database(db_path):
    """如果数据库不存在，则创建表并建立索引"""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute(CREATE_TABLE_SQL)
    c.execute(CREATE_INDEX_SQL)
    conn.commit()
    conn.close()


# ---------- 主界面 ----------
class MainScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(10))

        # 标题（显示当前账单名）
        self.title_label = Label(
            text='我的记账本',
            size_hint_y=0.12,
            font_size=24,
            halign='center'
        )
        layout.add_widget(self.title_label)

        # 账单选择下拉列表（仅显示名称，无后缀）
        bill_select_layout = BoxLayout(size_hint_y=0.08, spacing=dp(5))
        bill_select_layout.add_widget(Label(text='选择账单：', size_hint_x=0.25, font_size=dp(14)))
        self.bill_spinner = Spinner(
            text='',
            values=[],
            size_hint_x=0.75,
            font_size=dp(12)
        )
        self.bill_spinner.bind(text=self.on_bill_change)
        bill_select_layout.add_widget(self.bill_spinner)
        layout.add_widget(bill_select_layout)

        # 三个操作按钮：新建、重命名、删除
        btn_layout = BoxLayout(size_hint_y=0.1, spacing=dp(10))
        new_btn = Button(
            text='新建',
            background_color=(0.2, 0.7, 0.2, 1),
            font_size=dp(14)
        )
        new_btn.bind(on_press=self.new_bill_popup)
        btn_layout.add_widget(new_btn)

        rename_btn = Button(
            text='重命名',
            background_color=(0.3, 0.6, 0.9, 1),
            font_size=dp(14)
        )
        rename_btn.bind(on_press=self.rename_bill_popup)
        btn_layout.add_widget(rename_btn)

        delete_btn = Button(
            text='删除',
            background_color=(0.9, 0.2, 0.2, 1),
            font_size=dp(14)
        )
        delete_btn.bind(on_press=self.delete_bill_popup)
        btn_layout.add_widget(delete_btn)

        layout.add_widget(btn_layout)

        # 功能按钮
        view_btn = Button(
            text='查看事件',
            background_color=(0.3, 0.6, 0.9, 1),
            size_hint_y=0.12
        )
        view_btn.bind(on_press=self.go_to_view_event)
        layout.add_widget(view_btn)

        record_btn = Button(
            text='记收支',
            background_color=(0.2, 0.8, 0.5, 1),
            size_hint_y=0.12
        )
        record_btn.bind(on_press=self.go_to_record)
        layout.add_widget(record_btn)

        import_btn = Button(
            text='导入 Excel',
            background_color=(0.5, 0.5, 0.5, 1),
            size_hint_y=0.12
        )
        import_btn.bind(on_press=self.on_import)
        layout.add_widget(import_btn)

        export_btn = Button(
            text='导出 Excel',
            background_color=(0.5, 0.5, 0.5, 1),
            size_hint_y=0.12
        )
        export_btn.bind(on_press=self.on_export)
        layout.add_widget(export_btn)

        exit_btn = Button(
            text='退出该应用',
            background_color=(0.9, 0.2, 0.2, 1),
            size_hint_y=0.12,
            font_size=dp(16)
        )
        exit_btn.bind(on_press=self.exit_application)
        layout.add_widget(exit_btn)

        self.add_widget(layout)

    def on_enter(self):
        """每次进入主界面时刷新账单列表和当前显示"""
        app = App.get_running_app()
        app.refresh_bill_list()
        # 更新下拉列表（显示无后缀名）
        display_names = [os.path.splitext(name)[0] for name in app.bill_files]
        self.bill_spinner.values = display_names
        # 设置当前选中项
        if app.current_bill:
            current_display = os.path.splitext(app.current_bill)[0]
            self.bill_spinner.text = current_display
            self.update_title(current_display)

    def update_title(self, display_name):
        """更新标题显示，格式：我的记账本 - 账单名"""
        self.title_label.text = f'我的记账本 - {display_name}'

    def on_bill_change(self, spinner, text):
        """当下拉列表选择变化时，根据显示名称找到对应的真实文件名并切换"""
        if not text:
            return
        app = App.get_running_app()
        # 根据显示名称找到对应的真实文件名（带后缀）
        for filename in app.bill_files:
            if os.path.splitext(filename)[0] == text:
                if app.current_bill != filename:
                    app.current_bill = filename
                    self.update_title(text)
                break

    def new_bill_popup(self, instance):
        """弹出新建账单的输入窗口"""
        content = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(10))
        content.add_widget(Label(text='请输入新账单名称：'))
        name_input = TextInput(hint_text='例如：2025年家庭账本', multiline=False)
        content.add_widget(name_input)

        btn_layout = BoxLayout(size_hint_y=0.3, spacing=dp(10))
        cancel_btn = Button(text='取消')
        ok_btn = Button(text='确定', background_color=(0.2, 0.7, 0.2, 1))
        btn_layout.add_widget(cancel_btn)
        btn_layout.add_widget(ok_btn)
        content.add_widget(btn_layout)

        popup = Popup(title='新建账单', content=content, size_hint=(0.8, 0.4))

        def on_ok(btn):
            name = name_input.text.strip()
            app = App.get_running_app()
            success, msg = app.create_new_bill(name)
            if success:
                # 刷新下拉列表
                display_names = [os.path.splitext(name)[0] for name in app.bill_files]
                self.bill_spinner.values = display_names
                current_display = os.path.splitext(app.current_bill)[0]
                self.bill_spinner.text = current_display
                self.update_title(current_display)
                popup.dismiss()
            else:
                name_input.text = ''
                name_input.hint_text = msg
        ok_btn.bind(on_press=on_ok)
        cancel_btn.bind(on_press=popup.dismiss)
        popup.open()

    def rename_bill_popup(self, instance):
        """弹出重命名账单的输入窗口"""
        app = App.get_running_app()
        if not app.current_bill:
            return

        current_display = os.path.splitext(app.current_bill)[0]

        content = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(10))
        content.add_widget(Label(text='请输入新账单名称：'))
        name_input = TextInput(text=current_display, multiline=False, hint_text='新名称')
        content.add_widget(name_input)

        btn_layout = BoxLayout(size_hint_y=0.3, spacing=dp(10))
        cancel_btn = Button(text='取消')
        ok_btn = Button(text='确定', background_color=(0.3, 0.6, 0.9, 1))
        btn_layout.add_widget(cancel_btn)
        btn_layout.add_widget(ok_btn)
        content.add_widget(btn_layout)

        popup = Popup(title='重命名账单', content=content, size_hint=(0.8, 0.4))

        def on_ok(btn):
            new_name = name_input.text.strip()
            if not new_name:
                name_input.hint_text = '名称不能为空'
                return
            app = App.get_running_app()
            success, msg = app.rename_current_bill(new_name)
            if success:
                # 刷新下拉列表
                display_names = [os.path.splitext(name)[0] for name in app.bill_files]
                self.bill_spinner.values = display_names
                current_display = os.path.splitext(app.current_bill)[0]
                self.bill_spinner.text = current_display
                self.update_title(current_display)
                popup.dismiss()
            else:
                name_input.text = ''
                name_input.hint_text = msg
        ok_btn.bind(on_press=on_ok)
        cancel_btn.bind(on_press=popup.dismiss)
        popup.open()

    def delete_bill_popup(self, instance):
        """弹出删除账单确认窗口，要求输入'删除'二字确认"""
        app = App.get_running_app()
        if not app.current_bill:
            return

        current_display = os.path.splitext(app.current_bill)[0]

        content = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(10))
        content.add_widget(Label(
            text=f'确定要删除账单“{current_display}”吗？\n删除后无法恢复。\n请在下方输入“删除”以确认：',
            halign='center'
        ))
        confirm_input = TextInput(multiline=False, hint_text='输入“删除”')
        content.add_widget(confirm_input)

        btn_layout = BoxLayout(size_hint_y=0.3, spacing=dp(10))
        cancel_btn = Button(text='取消')
        ok_btn = Button(text='确认删除', background_color=(0.9, 0.2, 0.2, 1))
        btn_layout.add_widget(cancel_btn)
        btn_layout.add_widget(ok_btn)
        content.add_widget(btn_layout)

        popup = Popup(title='删除账单', content=content, size_hint=(0.8, 0.5))

        def on_ok(btn):
            if confirm_input.text.strip() == '删除':
                app = App.get_running_app()
                success, msg = app.delete_current_bill()
                if success:
                    # 刷新下拉列表
                    display_names = [os.path.splitext(name)[0] for name in app.bill_files]
                    self.bill_spinner.values = display_names
                    if app.current_bill:
                        current_display = os.path.splitext(app.current_bill)[0]
                        self.bill_spinner.text = current_display
                        self.update_title(current_display)
                    else:
                        self.bill_spinner.text = ''
                        self.title_label.text = '我的记账本'
                    popup.dismiss()
                else:
                    confirm_input.text = ''
                    confirm_input.hint_text = msg
            else:
                confirm_input.text = ''
                confirm_input.hint_text = '请输入“删除”'
        ok_btn.bind(on_press=on_ok)
        cancel_btn.bind(on_press=popup.dismiss)
        popup.open()

    def go_to_view_event(self, instance):
        self.manager.current = 'view_event'

    def go_to_record(self, instance):
        self.manager.current = 'record'

    def on_import(self, instance):
        self.manager.current = 'import'

    def on_export(self, instance):
        self.manager.current = 'export'

    def exit_application(self, instance):
        App.get_running_app().stop()

# ---------- 记收支页面（使用SQLite）----------
class RecordScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # 主布局：垂直方向
        main_layout = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(8))

        # 1. 标题
        title = Label(
            text='记录收支',
            size_hint_y=0.08,
            font_size=22,
            halign='center'
        )
        main_layout.add_widget(title)

        # 2. 交易日期选择
        date_layout = BoxLayout(size_hint_y=0.1, spacing=dp(5))
        date_layout.add_widget(Label(text='交易日期：', size_hint_x=0.2))

        date_sub_layout = BoxLayout(size_hint_x=0.8, spacing=dp(3))

        now = datetime.datetime.now()
        current_year = now.year if START_YEAR <= now.year <= END_YEAR else START_YEAR
        current_month = now.month
        current_day = now.day

        self.year_spinner = Spinner(
            text=str(current_year),
            values=YEAR_LIST,
            size_hint_x=0.33,
            font_size=dp(12)
        )
        date_sub_layout.add_widget(self.year_spinner)

        months = [f"{m:02d}" for m in range(1, 13)]
        self.month_spinner = Spinner(
            text=f"{current_month:02d}",
            values=months,
            size_hint_x=0.33,
            font_size=dp(12)
        )
        date_sub_layout.add_widget(self.month_spinner)

        days = [f"{d:02d}" for d in range(1, 32)]
        self.day_spinner = Spinner(
            text=f"{current_day:02d}",
            values=days,
            size_hint_x=0.33,
            font_size=dp(12)
        )
        date_sub_layout.add_widget(self.day_spinner)

        date_layout.add_widget(date_sub_layout)
        main_layout.add_widget(date_layout)

        # 3. 收支类型选择
        type_layout = BoxLayout(size_hint_y=0.09, spacing=dp(5))
        type_layout.add_widget(Label(text='收支类型：', size_hint_x=0.2))
        self.type_spinner = Spinner(
            text='收入',
            values=('收入', '支出'),
            size_hint_x=0.8,
            font_size=dp(12)
        )
        type_layout.add_widget(self.type_spinner)
        main_layout.add_widget(type_layout)

        # 4. 金额输入
        amount_layout = BoxLayout(size_hint_y=0.09, spacing=dp(5))
        amount_layout.add_widget(Label(text='金额（元）：', size_hint_x=0.2))
        self.amount_input = TextInput(
            hint_text='请输入数字，如 100.50',
            input_filter='float',
            size_hint_x=0.8,
            font_size=dp(12)
        )
        amount_layout.add_widget(self.amount_input)
        main_layout.add_widget(amount_layout)

        # 5. 事件（因由）输入
        event_layout = BoxLayout(size_hint_y=0.09, spacing=dp(5))
        event_layout.add_widget(Label(text='因由：', size_hint_x=0.2))
        self.event_input = TextInput(
            hint_text='如 工资发放、餐饮消费',
            size_hint_x=0.8,
            font_size=dp(12)
        )
        event_layout.add_widget(self.event_input)
        main_layout.add_widget(event_layout)

        # 6. 相关方输入
        party_layout = BoxLayout(size_hint_y=0.09, spacing=dp(5))
        party_layout.add_widget(Label(text='相关方：', size_hint_x=0.2))
        self.party_input = TextInput(
            hint_text='如 公司A、超市B',
            size_hint_x=0.8,
            font_size=dp(12)
        )
        party_layout.add_widget(self.party_input)
        main_layout.add_widget(party_layout)

        # 7. 交易平台输入
        platform_layout = BoxLayout(size_hint_y=0.09, spacing=dp(5))
        platform_layout.add_widget(Label(text='交易平台：', size_hint_x=0.2))
        self.platform_input = TextInput(
            hint_text='如 微信支付、支付宝、招商银行',
            size_hint_x=0.8,
            font_size=dp(12)
        )
        platform_layout.add_widget(self.platform_input)
        main_layout.add_widget(platform_layout)

        # 8. 其他信息输入
        other_layout = BoxLayout(size_hint_y=0.09, spacing=dp(5))
        other_layout.add_widget(Label(text='其他：', size_hint_x=0.2))
        self.other_input = TextInput(
            hint_text='补充说明信息',
            size_hint_x=0.8,
            font_size=dp(12)
        )
        other_layout.add_widget(self.other_input)
        main_layout.add_widget(other_layout)

        # 9. 占位布局 (可以适当调整高度)
        placeholder = BoxLayout(size_hint_y=0.1)
        main_layout.add_widget(placeholder)

        # 10. 保存按钮和保存并退出按钮放在一行
        btn_row = BoxLayout(size_hint_y=0.09, spacing=dp(5))
        save_only_btn = Button(
            text='保存',
            background_color=(0.2, 0.7, 0.2, 1),
            font_size=dp(14)
        )
        save_only_btn.bind(on_press=self.save_only)
        btn_row.add_widget(save_only_btn)

        save_and_exit_btn = Button(
            text='保存并退出',
            background_color=(0.2, 0.7, 0.2, 1),
            font_size=dp(14)
        )
        save_and_exit_btn.bind(on_press=self.save_and_exit)
        btn_row.add_widget(save_and_exit_btn)

        main_layout.add_widget(btn_row)

        # 11. 返回主界面按钮
        back_btn = Button(
            text='返回主界面',
            background_color=(0.7, 0.7, 0.7, 1),
            size_hint_y=0.09,
            font_size=dp(14)
        )
        back_btn.bind(on_press=self.go_back_to_main)
        main_layout.add_widget(back_btn)

        # 12. 提示标签
        self.tip_label = Label(
            text='',
            size_hint_y=0.05,
            halign='center',
            color=(1, 0, 0, 1),
            font_size=dp(12)
        )
        main_layout.add_widget(self.tip_label)

        self.add_widget(main_layout)

    def save_and_exit(self, instance):
        """保存收支数据到当前账单的SQLite数据库并返回主界面"""
        amount_text = self.amount_input.text.strip()
        if not amount_text:
            self.tip_label.text = '请输入金额！'
            return
        try:
            amount = float(amount_text)
            if amount <= 0:
                self.tip_label.text = '金额必须大于0！'
                return
        except ValueError:
            self.tip_label.text = '金额格式错误！'
            return

        try:
            year = self.year_spinner.text
            month = self.month_spinner.text
            day = self.day_spinner.text
            datetime.datetime(int(year), int(month), int(day))
            date_str = f"{year}-{month}-{day}"
        except ValueError as e:
            self.tip_label.text = f'无效日期：{e}'
            return

        trade_type = self.type_spinner.text
        event = self.event_input.text.strip() or ''
        party = self.party_input.text.strip() or ''
        platform = self.platform_input.text.strip() or ''
        other = self.other_input.text.strip() or ''

        app = App.get_running_app()
        db_path = app.get_current_db_path()
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        try:
            c.execute('''
                INSERT INTO records (date, type, amount, event, party, platform, other)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (date_str, trade_type, amount, event, party, platform, other))
            conn.commit()
            self.tip_label.text = ''
            self.amount_input.text = ''
            self.event_input.text = ''
            self.party_input.text = ''
            self.platform_input.text = ''
            self.other_input.text = ''
            self.manager.current = 'main'
        except Exception as e:
            self.tip_label.text = f'保存失败：{str(e)}'
        finally:
            conn.close()

    def _save_record(self):
        """保存记录到数据库，返回 (成功标志, 错误信息或None)"""
        amount_text = self.amount_input.text.strip()
        if not amount_text:
            return False, '请输入金额！'
        try:
            amount = float(amount_text)
            if amount <= 0:
                return False, '金额必须大于0！'
        except ValueError:
            return False, '金额格式错误！'

        try:
            year = self.year_spinner.text
            month = self.month_spinner.text
            day = self.day_spinner.text
            datetime.datetime(int(year), int(month), int(day))
            date_str = f"{year}-{month}-{day}"
        except ValueError as e:
            return False, f'无效日期：{e}'

        trade_type = self.type_spinner.text
        event = self.event_input.text.strip() or ''
        party = self.party_input.text.strip() or ''
        platform = self.platform_input.text.strip() or ''
        other = self.other_input.text.strip() or ''

        app = App.get_running_app()
        db_path = app.get_current_db_path()
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        try:
            c.execute('''
                INSERT INTO records (date, type, amount, event, party, platform, other)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (date_str, trade_type, amount, event, party, platform, other))
            conn.commit()
            # 清空输入字段（保留日期和类型选择不变）
            self.amount_input.text = ''
            self.event_input.text = ''
            self.party_input.text = ''
            self.platform_input.text = ''
            self.other_input.text = ''
            return True, None
        except Exception as e:
            return False, f'保存失败：{str(e)}'
        finally:
            conn.close()

    def save_only(self, instance):
        """保存记录但不退出"""
        success, msg = self._save_record()
        if success:
            self.tip_label.text = '保存成功'
            self.tip_label.color = (0, 0.7, 0, 1)
        else:
            self.tip_label.text = msg
            self.tip_label.color = (1, 0, 0, 1)

    def save_and_exit(self, instance):
        """保存记录并返回主界面"""
        success, msg = self._save_record()
        if success:
            self.tip_label.text = ''
            self.manager.current = 'main'
        else:
            self.tip_label.text = msg
            self.tip_label.color = (1, 0, 0, 1)

    def go_back_to_main(self, instance):
        self.tip_label.text = ''
        self.manager.current = 'main'


# ---------- 查看事件界面（使用SQLite）----------
class ViewEventScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        main_layout = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(5))

        back_btn = Button(
            text='返回主界面',
            size_hint_y=None,
            height=dp(35),
            background_color=(0.8, 0.8, 0.8, 1)
        )
        back_btn.bind(on_press=self.go_back)
        main_layout.add_widget(back_btn)

        self.type_spinner = Spinner(
            text='天',
            values=('天', '月', '年', '时间段'),
            size_hint_y=None,
            height=dp(35)
        )
        self.type_spinner.bind(text=self.on_type_change)
        main_layout.add_widget(self.type_spinner)

        self.date_container = BoxLayout(orientation='vertical', size_hint_y=None, spacing=dp(2))
        main_layout.add_widget(self.date_container)

        self.income_type_spinner = Spinner(
            text='全部',
            values=('全部', '收入', '支出'),
            size_hint_y=None,
            height=dp(35)
        )
        main_layout.add_widget(self.income_type_spinner)

        query_btn = Button(
            text='查询',
            size_hint_y=None,
            height=dp(35),
            background_color=(0.3, 0.6, 0.9, 1)
        )
        query_btn.bind(on_press=self.query_events)
        main_layout.add_widget(query_btn)

        from kivy.uix.scrollview import ScrollView
        scroll_view = ScrollView(
            size_hint=(1, 1),
            do_scroll_x=False,
            do_scroll_y=True
        )
        self.result_label = Label(
            text='请选择查询条件并点击查询',
            size_hint_y=None,
            size_hint_x=1,
            halign='left',
            valign='top',
            padding=dp(5)
        )
        self.result_label.bind(
            texture_size=lambda instance, value: setattr(instance, 'height', value[1])
        )
        self.result_label.bind(
            width=lambda instance, value: setattr(instance, 'text_size', (value - dp(10), None))
        )
        scroll_view.add_widget(self.result_label)
        main_layout.add_widget(scroll_view)

        self.add_widget(main_layout)

        self.date_spinners = {}
        self.update_date_controls('天')

    def go_back(self, instance):
        self.manager.current = 'main'

    def on_type_change(self, instance, text):
        self.update_date_controls(text)

    def update_date_controls(self, query_type):
        self.date_container.clear_widgets()
        self.date_spinners.clear()

        now = datetime.datetime.now()
        current_year = now.year if START_YEAR <= now.year <= END_YEAR else START_YEAR
        current_month = now.month
        current_day = now.day

        spinner_style = {
            'size_hint_y': None,
            'height': dp(30),
            'font_size': dp(12)
        }

        if query_type == '天':
            h_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30), spacing=dp(3))
            year_sp = Spinner(text=str(current_year), values=YEAR_LIST, size_hint_x=0.33, **spinner_style)
            h_layout.add_widget(year_sp)
            self.date_spinners['year'] = year_sp

            months = [f"{m:02d}" for m in range(1, 13)]
            month_sp = Spinner(text=f"{current_month:02d}", values=months, size_hint_x=0.33, **spinner_style)
            h_layout.add_widget(month_sp)
            self.date_spinners['month'] = month_sp

            days = [f"{d:02d}" for d in range(1, 32)]
            day_sp = Spinner(text=f"{current_day:02d}", values=days, size_hint_x=0.33, **spinner_style)
            h_layout.add_widget(day_sp)
            self.date_spinners['day'] = day_sp

            self.date_container.add_widget(h_layout)

        elif query_type == '月':
            h_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30), spacing=dp(3))
            year_sp = Spinner(text=str(current_year), values=YEAR_LIST, size_hint_x=0.5, **spinner_style)
            h_layout.add_widget(year_sp)
            self.date_spinners['year'] = year_sp

            months = [f"{m:02d}" for m in range(1, 13)]
            month_sp = Spinner(text=f"{current_month:02d}", values=months, size_hint_x=0.5, **spinner_style)
            h_layout.add_widget(month_sp)
            self.date_spinners['month'] = month_sp

            self.date_container.add_widget(h_layout)

        elif query_type == '年':
            h_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30))
            year_sp = Spinner(text=str(current_year), values=YEAR_LIST, size_hint_x=1, **spinner_style)
            h_layout.add_widget(year_sp)
            self.date_spinners['year'] = year_sp

            self.date_container.add_widget(h_layout)

        elif query_type == '时间段':
            start_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30), spacing=dp(2))
            start_layout.add_widget(Label(text='从', size_hint_x=0.08, font_size=dp(12), halign='center'))
            start_year = Spinner(text=str(current_year), values=YEAR_LIST, size_hint_x=0.28, **spinner_style)
            start_layout.add_widget(start_year)
            self.date_spinners['start_year'] = start_year

            months = [f"{m:02d}" for m in range(1, 13)]
            start_month = Spinner(text=f"{current_month:02d}", values=months, size_hint_x=0.28, **spinner_style)
            start_layout.add_widget(start_month)
            self.date_spinners['start_month'] = start_month

            days = [f"{d:02d}" for d in range(1, 32)]
            start_day = Spinner(text=f"{current_day:02d}", values=days, size_hint_x=0.28, **spinner_style)
            start_layout.add_widget(start_day)
            self.date_spinners['start_day'] = start_day

            self.date_container.add_widget(start_layout)

            end_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30), spacing=dp(2))
            end_layout.add_widget(Label(text='到', size_hint_x=0.08, font_size=dp(12), halign='center'))
            end_year = Spinner(text=str(current_year), values=YEAR_LIST, size_hint_x=0.28, **spinner_style)
            end_layout.add_widget(end_year)
            self.date_spinners['end_year'] = end_year

            end_month = Spinner(text=f"{current_month:02d}", values=months, size_hint_x=0.28, **spinner_style)
            end_layout.add_widget(end_month)
            self.date_spinners['end_month'] = end_month

            end_day = Spinner(text=f"{current_day:02d}", values=days, size_hint_x=0.28, **spinner_style)
            end_layout.add_widget(end_day)
            self.date_spinners['end_day'] = end_day

            self.date_container.add_widget(end_layout)

    def query_events(self, instance):
        query_type = self.type_spinner.text
        income_filter = self.income_type_spinner.text

        try:
            if query_type == '天':
                year = self.date_spinners['year'].text
                month = self.date_spinners['month'].text
                day = self.date_spinners['day'].text
                start_date = f"{year}-{month}-{day}"
                end_date = start_date
            elif query_type == '月':
                year = self.date_spinners['year'].text
                month = self.date_spinners['month'].text
                start_date = f"{year}-{month}-01"
                if month == '12':
                    end_date = f"{int(year)+1}-01-01"
                else:
                    end_date = f"{year}-{int(month)+1:02d}-01"
            elif query_type == '年':
                year = self.date_spinners['year'].text
                start_date = f"{year}-01-01"
                end_date = f"{int(year)+1}-01-01"
            elif query_type == '时间段':
                start_year = self.date_spinners['start_year'].text
                start_month = self.date_spinners['start_month'].text
                start_day = self.date_spinners['start_day'].text
                end_year = self.date_spinners['end_year'].text
                end_month = self.date_spinners['end_month'].text
                end_day = self.date_spinners['end_day'].text
                start_date = f"{start_year}-{start_month}-{start_day}"
                end_date = f"{end_year}-{end_month}-{end_day}"
            else:
                self.result_label.text = "未知查询类型"
                return
        except KeyError as e:
            self.result_label.text = f"日期控件未初始化：{e}"
            return

        app = App.get_running_app()
        db_path = app.get_current_db_path()
        if not os.path.exists(db_path):
            self.result_label.text = "当前账单数据库不存在，请先记录一笔收支。"
            return

        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        try:
            sql = "SELECT date, type, amount, event, party, platform, other FROM records"
            conditions = []
            params = []

            if query_type in ('天', '时间段'):
                conditions.append("date BETWEEN ? AND ?")
                params.extend([start_date, end_date])
            elif query_type in ('月', '年'):
                conditions.append("date >= ? AND date < ?")
                params.extend([start_date, end_date])

            if income_filter == '收入':
                conditions.append("type = '收入'")
            elif income_filter == '支出':
                conditions.append("type = '支出'")

            if conditions:
                sql += " WHERE " + " AND ".join(conditions)

            sql += " ORDER BY date"

            c.execute(sql, params)
            rows = c.fetchall()

            if not rows:
                self.result_label.text = "没有找到符合条件的记录。"
                return

            lines = []
            total_income = 0.0
            total_expense = 0.0

            for row in rows:
                date_val, typ, amount, event, party, platform, other = row
                date_val = date_val or ''
                typ = typ or ''
                amount = amount or 0.0
                event = event or ''
                party = party or ''
                platform = platform or ''
                other = other or ''

                if typ == '收入':
                    total_income += amount
                elif typ == '支出':
                    total_expense += amount

                line = f'您于"{date_val}"，"{typ}"{amount:.2f}元，因由："{event}"，相关方："{party}"，交易平台："{platform}"，其他："{other}"'
                lines.append(line)

            lines.append("")
            lines.append(f"总收入：{total_income:.2f} 元")
            lines.append(f"总支出：{total_expense:.2f} 元")
            lines.append(f"结余：{total_income - total_expense:.2f} 元")

            self.result_label.text = "\n".join(lines)

        except Exception as e:
            self.result_label.text = f"查询出错：{str(e)}"
        finally:
            conn.close()


# ---------- 导入页面（从Excel导入到SQLite）----------
class ImportScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.selected_file = None

        main_layout = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(10))

        title = Label(text='导入 Excel', size_hint_y=0.1, font_size=22, halign='center')
        main_layout.add_widget(title)

        # 导入模式
        mode_layout = BoxLayout(size_hint_y=0.08, spacing=dp(5))
        mode_layout.add_widget(Label(text='导入模式：', size_hint_x=0.25, font_size=dp(14)))
        self.mode_spinner = Spinner(
            text='导入成为新账单',
            values=('导入成为新账单', '与指定账单合并为新账单', '并入指定账单'),
            size_hint_x=0.75,
            font_size=dp(12)
        )
        self.mode_spinner.bind(text=self.on_mode_change)
        mode_layout.add_widget(self.mode_spinner)
        main_layout.add_widget(mode_layout)

        # 选择外部文件
        file_layout = BoxLayout(size_hint_y=0.08, spacing=dp(5))
        file_layout.add_widget(Label(text='外部文件：', size_hint_x=0.25, font_size=dp(14)))
        self.file_path_label = Label(text='未选择', size_hint_x=0.5, font_size=dp(12), color=(0.5,0.5,0.5,1))
        file_layout.add_widget(self.file_path_label)
        choose_file_btn = Button(text='浏览', size_hint_x=0.25, font_size=dp(12))
        choose_file_btn.bind(on_press=self.choose_file)
        file_layout.add_widget(choose_file_btn)
        main_layout.add_widget(file_layout)

        # 目标账单选择
        self.target_layout = BoxLayout(size_hint_y=0.08, spacing=dp(5))
        self.target_layout.add_widget(Label(text='目标账单：', size_hint_x=0.25, font_size=dp(14)))
        self.target_spinner = Spinner(
            text='',
            values=[],
            size_hint_x=0.75,
            font_size=dp(12)
        )
        self.target_layout.add_widget(self.target_spinner)
        main_layout.add_widget(self.target_layout)

        # 新账单名输入
        self.new_name_layout = BoxLayout(size_hint_y=0.08, spacing=dp(5))
        self.new_name_layout.add_widget(Label(text='新账单名：', size_hint_x=0.25, font_size=dp(14)))
        self.new_name_input = TextInput(
            hint_text='例如：合并账本',
            size_hint_x=0.75,
            font_size=dp(12),
            multiline=False
        )
        self.new_name_layout.add_widget(self.new_name_input)
        main_layout.add_widget(self.new_name_layout)

        main_layout.add_widget(BoxLayout(size_hint_y=0.1))

        # 执行导入按钮
        import_btn = Button(
            text='执行导入',
            size_hint_y=0.08,
            background_color=(0.2, 0.7, 0.2, 1),
            font_size=dp(14)
        )
        import_btn.bind(on_press=self.do_import)
        main_layout.add_widget(import_btn)

        # 返回按钮
        back_btn = Button(
            text='返回主界面',
            size_hint_y=0.08,
            background_color=(0.7, 0.7, 0.7, 1),
            font_size=dp(14)
        )
        back_btn.bind(on_press=self.go_back)
        main_layout.add_widget(back_btn)

        # 提示标签
        self.tip_label = Label(
            text='',
            size_hint_y=0.05,
            halign='center',
            color=(1,0,0,1),
            font_size=dp(12)
        )
        main_layout.add_widget(self.tip_label)

        self.add_widget(main_layout)
        self.on_mode_change(None, self.mode_spinner.text)

    def on_enter(self):
        app = App.get_running_app()
        display_names = [os.path.splitext(name)[0] for name in app.bill_files]
        self.target_spinner.values = display_names
        if app.current_bill:
            self.target_spinner.text = os.path.splitext(app.current_bill)[0]

    def on_mode_change(self, instance, mode):
        if mode == '导入成为新账单':
            self.target_layout.disabled = True
            self.target_layout.opacity = 0.3
            self.new_name_layout.disabled = False
            self.new_name_layout.opacity = 1
        elif mode == '与指定账单合并为新账单':
            self.target_layout.disabled = False
            self.target_layout.opacity = 1
            self.new_name_layout.disabled = False
            self.new_name_layout.opacity = 1
        elif mode == '并入指定账单':
            self.target_layout.disabled = False
            self.target_layout.opacity = 1
            self.new_name_layout.disabled = True
            self.new_name_layout.opacity = 0.3

    def choose_file(self, instance):
        filechooser.open_file(
            on_selection=self.on_file_selected,
            filters=["*.xlsx"]
        )

    def on_file_selected(self, selection):
        if selection:
            self.selected_file = selection[0]
            self.file_path_label.text = os.path.basename(self.selected_file)
            self.file_path_label.color = (0,0,0,1)
        else:
            self.selected_file = None
            self.file_path_label.text = '未选择'
            self.file_path_label.color = (0.5,0.5,0.5,1)

    def _read_excel_to_records(self, filepath):
        """读取Excel文件，返回记录列表（每个记录为字典），并验证必要列"""
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = rows[0]
        # 期望的列名映射（允许大小写和空格差异）
        expected = ['时间', '收支', '金额', '事件', '相关方', '交易平台', '其他']
        col_indices = {}
        for i, h in enumerate(headers):
            if h:
                h_norm = h.strip()
                for exp in expected:
                    if h_norm == exp:
                        col_indices[exp] = i
                        break
        # 确保必要列存在
        required = ['时间', '收支', '金额']
        for r in required:
            if r not in col_indices:
                raise ValueError(f"Excel中缺少必要列：{r}")

        records = []
        for row in rows[1:]:
            if not any(row):  # 跳过全空行
                continue
            date_cell = row[col_indices['时间']] if col_indices['时间'] < len(row) else None
            if isinstance(date_cell, datetime.datetime):
                date_str = date_cell.strftime('%Y-%m-%d')
            else:
                date_str = str(date_cell).strip() if date_cell else ''
            if not date_str:
                continue  # 无日期则跳过
            type_cell = row[col_indices['收支']] if col_indices['收支'] < len(row) else ''
            type_str = str(type_cell).strip() if type_cell else ''
            if type_str not in ('收入', '支出'):
                continue  # 跳过无效类型

            amount_cell = row[col_indices['金额']] if col_indices['金额'] < len(row) else 0
            try:
                amount = float(amount_cell) if amount_cell not in (None, '') else 0.0
            except:
                amount = 0.0

            # 其他字段
            def get_val(idx):
                if idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val else ''
                return ''
            event = get_val(col_indices.get('事件', -1))
            party = get_val(col_indices.get('相关方', -1))
            platform = get_val(col_indices.get('交易平台', -1))
            other = get_val(col_indices.get('其他', -1))

            records.append({
                'date': date_str,
                'type': type_str,
                'amount': amount,
                'event': event,
                'party': party,
                'platform': platform,
                'other': other
            })
        return records

    def do_import(self, instance):
        mode = self.mode_spinner.text
        app = App.get_running_app()

        if not self.selected_file:
            self.tip_label.text = '请先选择外部 Excel 文件'
            return

        try:
            records = self._read_excel_to_records(self.selected_file)
        except Exception as e:
            self.tip_label.text = f'读取外部文件失败：{str(e)}'
            return

        if not records:
            self.tip_label.text = 'Excel中没有有效数据'
            return

        try:
            if mode == '导入成为新账单':
                new_name = self.new_name_input.text.strip()
                if not new_name:
                    self.tip_label.text = '请输入新账单名称'
                    return
                # 创建新数据库文件
                filename = new_name + '.db'
                filepath = os.path.join(app.bills_dir, filename)
                if os.path.exists(filepath):
                    self.tip_label.text = '该账单名称已存在，请换一个'
                    return
                # 初始化数据库并插入数据
                init_database(filepath)
                conn = sqlite3.connect(filepath)
                c = conn.cursor()
                for rec in records:
                    c.execute('''
                        INSERT INTO records (date, type, amount, event, party, platform, other)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (rec['date'], rec['type'], rec['amount'], rec['event'],
                          rec['party'], rec['platform'], rec['other']))
                conn.commit()
                conn.close()
                app.refresh_bill_list()
                app.current_bill = filename
                self.tip_label.text = f'导入成功，已切换到新账单“{new_name}”'

            elif mode == '与指定账单合并为新账单':
                target_display = self.target_spinner.text
                if not target_display:
                    self.tip_label.text = '请选择目标账单'
                    return
                new_name = self.new_name_input.text.strip()
                if not new_name:
                    self.tip_label.text = '请输入新账单名称'
                    return
                target_file = None
                for f in app.bill_files:
                    if os.path.splitext(f)[0] == target_display:
                        target_file = f
                        break
                if not target_file:
                    self.tip_label.text = '目标账单不存在'
                    return

                # 读取目标数据库的所有记录
                target_path = os.path.join(app.bills_dir, target_file)
                conn_src = sqlite3.connect(target_path)
                c_src = conn_src.cursor()
                c_src.execute("SELECT date, type, amount, event, party, platform, other FROM records")
                src_rows = c_src.fetchall()
                conn_src.close()

                # 创建新数据库
                new_filename = new_name + '.db'
                new_path = os.path.join(app.bills_dir, new_filename)
                if os.path.exists(new_path):
                    self.tip_label.text = '该账单名称已存在，请换一个'
                    return
                init_database(new_path)
                conn_new = sqlite3.connect(new_path)
                c_new = conn_new.cursor()

                # 插入源数据
                for row in src_rows:
                    c_new.execute('''
                        INSERT INTO records (date, type, amount, event, party, platform, other)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', row)
                # 插入导入数据
                for rec in records:
                    c_new.execute('''
                        INSERT INTO records (date, type, amount, event, party, platform, other)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (rec['date'], rec['type'], rec['amount'], rec['event'],
                          rec['party'], rec['platform'], rec['other']))
                conn_new.commit()
                conn_new.close()
                app.refresh_bill_list()
                app.current_bill = new_filename
                self.tip_label.text = f'合并成功，已切换到新账单“{new_name}”'

            elif mode == '并入指定账单':
                target_display = self.target_spinner.text
                if not target_display:
                    self.tip_label.text = '请选择目标账单'
                    return
                target_file = None
                for f in app.bill_files:
                    if os.path.splitext(f)[0] == target_display:
                        target_file = f
                        break
                if not target_file:
                    self.tip_label.text = '目标账单不存在'
                    return

                target_path = os.path.join(app.bills_dir, target_file)
                conn = sqlite3.connect(target_path)
                c = conn.cursor()
                for rec in records:
                    c.execute('''
                        INSERT INTO records (date, type, amount, event, party, platform, other)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (rec['date'], rec['type'], rec['amount'], rec['event'],
                          rec['party'], rec['platform'], rec['other']))
                conn.commit()
                conn.close()
                self.tip_label.text = f'并入成功，账单“{target_display}”已更新'

            # 清空选择
            self.selected_file = None
            self.file_path_label.text = '未选择'
            self.file_path_label.color = (0.5,0.5,0.5,1)
            self.tip_label.color = (0,0.7,0,1)

        except Exception as e:
            self.tip_label.text = f'导入失败：{str(e)}'
            self.tip_label.color = (1,0,0,1)

    def go_back(self, instance):
        self.manager.current = 'main'


# ---------- 导出页面（从SQLite导出到Excel）----------
class ExportScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        main_layout = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(10))

        title = Label(text='导出账单副本', size_hint_y=0.1, font_size=22, halign='center')
        main_layout.add_widget(title)

        # 查询类型选择
        type_layout = BoxLayout(size_hint_y=0.08, spacing=dp(5))
        type_layout.add_widget(Label(text='查询类型：', size_hint_x=0.25, font_size=dp(14)))
        self.type_spinner = Spinner(
            text='天',
            values=('天', '月', '年', '时间段'),
            size_hint_x=0.75,
            font_size=dp(12)
        )
        self.type_spinner.bind(text=self.on_type_change)
        type_layout.add_widget(self.type_spinner)
        main_layout.add_widget(type_layout)

        # 动态日期容器
        self.date_container = BoxLayout(orientation='vertical', size_hint_y=None, spacing=dp(2))
        main_layout.add_widget(self.date_container)

        # 收支类型选择
        income_layout = BoxLayout(size_hint_y=0.08, spacing=dp(5))
        income_layout.add_widget(Label(text='收支类型：', size_hint_x=0.25, font_size=dp(14)))
        self.income_spinner = Spinner(
            text='全部',
            values=('全部', '收入', '支出'),
            size_hint_x=0.75,
            font_size=dp(12)
        )
        income_layout.add_widget(self.income_spinner)
        main_layout.add_widget(income_layout)

        # 目标文件夹选择
        folder_layout = BoxLayout(size_hint_y=0.08, spacing=dp(5))
        folder_layout.add_widget(Label(text='目标文件夹：', size_hint_x=0.25, font_size=dp(14)))
        self.folder_label = Label(
            text='未选择',
            size_hint_x=0.5,
            font_size=dp(12),
            color=(0.5, 0.5, 0.5, 1),
            shorten=True,
            shorten_from='right'
        )
        folder_layout.add_widget(self.folder_label)
        browse_btn = Button(text='浏览', size_hint_x=0.25, font_size=dp(12))
        browse_btn.bind(on_press=self.choose_folder)
        folder_layout.add_widget(browse_btn)
        main_layout.add_widget(folder_layout)

        # 文件名输入
        name_layout = BoxLayout(size_hint_y=0.08, spacing=dp(5))
        name_layout.add_widget(Label(text='文件名：', size_hint_x=0.25, font_size=dp(14)))
        self.filename_input = TextInput(
            text='',
            size_hint_x=0.75,
            font_size=dp(12),
            multiline=False,
            hint_text='例如：我的账单_导出'
        )
        name_layout.add_widget(self.filename_input)
        main_layout.add_widget(name_layout)

        main_layout.add_widget(BoxLayout(size_hint_y=0.05))

        # 导出按钮
        export_btn = Button(
            text='执行导出',
            size_hint_y=0.1,
            background_color=(0.2, 0.7, 0.2, 1),
            font_size=dp(16)
        )
        export_btn.bind(on_press=self.do_export)
        main_layout.add_widget(export_btn)

        # 返回主界面按钮
        back_btn = Button(
            text='返回主界面',
            size_hint_y=0.08,
            background_color=(0.7, 0.7, 0.7, 1),
            font_size=dp(14)
        )
        back_btn.bind(on_press=self.go_back)
        main_layout.add_widget(back_btn)

        # 提示标签
        self.tip_label = Label(
            text='',
            size_hint_y=0.05,
            halign='center',
            color=(1,0,0,1),
            font_size=dp(12)
        )
        main_layout.add_widget(self.tip_label)

        self.add_widget(main_layout)

        self.date_spinners = {}
        self.update_date_controls('天')
        self.selected_folder = None

    def on_enter(self):
        app = App.get_running_app()
        if app.current_bill:
            default_name = os.path.splitext(app.current_bill)[0]
            self.filename_input.text = default_name
        else:
            self.filename_input.text = ''

    def on_type_change(self, instance, text):
        self.update_date_controls(text)

    def update_date_controls(self, query_type):
        self.date_container.clear_widgets()
        self.date_spinners.clear()

        now = datetime.datetime.now()
        current_year = now.year if START_YEAR <= now.year <= END_YEAR else START_YEAR
        current_month = now.month
        current_day = now.day

        spinner_style = {
            'size_hint_y': None,
            'height': dp(30),
            'font_size': dp(12)
        }

        if query_type == '天':
            h_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30), spacing=dp(3))
            year_sp = Spinner(text=str(current_year), values=YEAR_LIST, size_hint_x=0.33, **spinner_style)
            h_layout.add_widget(year_sp)
            self.date_spinners['year'] = year_sp

            months = [f"{m:02d}" for m in range(1, 13)]
            month_sp = Spinner(text=f"{current_month:02d}", values=months, size_hint_x=0.33, **spinner_style)
            h_layout.add_widget(month_sp)
            self.date_spinners['month'] = month_sp

            days = [f"{d:02d}" for d in range(1, 32)]
            day_sp = Spinner(text=f"{current_day:02d}", values=days, size_hint_x=0.33, **spinner_style)
            h_layout.add_widget(day_sp)
            self.date_spinners['day'] = day_sp

            self.date_container.add_widget(h_layout)

        elif query_type == '月':
            h_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30), spacing=dp(3))
            year_sp = Spinner(text=str(current_year), values=YEAR_LIST, size_hint_x=0.5, **spinner_style)
            h_layout.add_widget(year_sp)
            self.date_spinners['year'] = year_sp

            months = [f"{m:02d}" for m in range(1, 13)]
            month_sp = Spinner(text=f"{current_month:02d}", values=months, size_hint_x=0.5, **spinner_style)
            h_layout.add_widget(month_sp)
            self.date_spinners['month'] = month_sp

            self.date_container.add_widget(h_layout)

        elif query_type == '年':
            h_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30))
            year_sp = Spinner(text=str(current_year), values=YEAR_LIST, size_hint_x=1, **spinner_style)
            h_layout.add_widget(year_sp)
            self.date_spinners['year'] = year_sp

            self.date_container.add_widget(h_layout)

        elif query_type == '时间段':
            start_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30), spacing=dp(2))
            start_layout.add_widget(Label(text='从', size_hint_x=0.08, font_size=dp(12), halign='center'))
            start_year = Spinner(text=str(current_year), values=YEAR_LIST, size_hint_x=0.28, **spinner_style)
            start_layout.add_widget(start_year)
            self.date_spinners['start_year'] = start_year

            months = [f"{m:02d}" for m in range(1, 13)]
            start_month = Spinner(text=f"{current_month:02d}", values=months, size_hint_x=0.28, **spinner_style)
            start_layout.add_widget(start_month)
            self.date_spinners['start_month'] = start_month

            days = [f"{d:02d}" for d in range(1, 32)]
            start_day = Spinner(text=f"{current_day:02d}", values=days, size_hint_x=0.28, **spinner_style)
            start_layout.add_widget(start_day)
            self.date_spinners['start_day'] = start_day

            self.date_container.add_widget(start_layout)

            end_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(30), spacing=dp(2))
            end_layout.add_widget(Label(text='到', size_hint_x=0.08, font_size=dp(12), halign='center'))
            end_year = Spinner(text=str(current_year), values=YEAR_LIST, size_hint_x=0.28, **spinner_style)
            end_layout.add_widget(end_year)
            self.date_spinners['end_year'] = end_year

            end_month = Spinner(text=f"{current_month:02d}", values=months, size_hint_x=0.28, **spinner_style)
            end_layout.add_widget(end_month)
            self.date_spinners['end_month'] = end_month

            end_day = Spinner(text=f"{current_day:02d}", values=days, size_hint_x=0.28, **spinner_style)
            end_layout.add_widget(end_day)
            self.date_spinners['end_day'] = end_day

            self.date_container.add_widget(end_layout)

    def choose_folder(self, instance):
        filechooser.choose_dir(on_selection=self.on_folder_selected)

    def on_folder_selected(self, selection):
        if selection:
            self.selected_folder = selection[0]
            self.folder_label.text = self.selected_folder
            self.folder_label.color = (0, 0, 0, 1)
        else:
            self.selected_folder = None
            self.folder_label.text = '未选择'
            self.folder_label.color = (0.5, 0.5, 0.5, 1)

    def do_export(self, instance):
        if not self.selected_folder:
            self.tip_label.text = '请先选择目标文件夹'
            self.tip_label.color = (1, 0, 0, 1)
            return

        filename = self.filename_input.text.strip()
        if not filename:
            self.tip_label.text = '请输入文件名'
            self.tip_label.color = (1, 0, 0, 1)
            return
        if not filename.lower().endswith('.xlsx'):
            filename += '.xlsx'

        full_path = os.path.join(self.selected_folder, filename)

        query_type = self.type_spinner.text
        income_filter = self.income_spinner.text

        try:
            if query_type == '天':
                year = self.date_spinners['year'].text
                month = self.date_spinners['month'].text
                day = self.date_spinners['day'].text
                start_date = f"{year}-{month}-{day}"
                end_date = start_date
            elif query_type == '月':
                year = self.date_spinners['year'].text
                month = self.date_spinners['month'].text
                start_date = f"{year}-{month}-01"
                if month == '12':
                    end_date = f"{int(year)+1}-01-01"
                else:
                    end_date = f"{year}-{int(month)+1:02d}-01"
            elif query_type == '年':
                year = self.date_spinners['year'].text
                start_date = f"{year}-01-01"
                end_date = f"{int(year)+1}-01-01"
            elif query_type == '时间段':
                start_year = self.date_spinners['start_year'].text
                start_month = self.date_spinners['start_month'].text
                start_day = self.date_spinners['start_day'].text
                end_year = self.date_spinners['end_year'].text
                end_month = self.date_spinners['end_month'].text
                end_day = self.date_spinners['end_day'].text
                start_date = f"{start_year}-{start_month}-{start_day}"
                end_date = f"{end_year}-{end_month}-{end_day}"
            else:
                self.tip_label.text = "未知查询类型"
                return
        except KeyError as e:
            self.tip_label.text = f"日期控件未初始化：{e}"
            return

        app = App.get_running_app()
        db_path = app.get_current_db_path()
        if not os.path.exists(db_path):
            self.tip_label.text = "当前账单数据库不存在，无数据可导出"
            return

        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        try:
            sql = "SELECT date, type, amount, event, party, platform, other FROM records"
            conditions = []
            params = []

            if query_type in ('天', '时间段'):
                conditions.append("date BETWEEN ? AND ?")
                params.extend([start_date, end_date])
            elif query_type in ('月', '年'):
                conditions.append("date >= ? AND date < ?")
                params.extend([start_date, end_date])

            if income_filter == '收入':
                conditions.append("type = '收入'")
            elif income_filter == '支出':
                conditions.append("type = '支出'")

            if conditions:
                sql += " WHERE " + " AND ".join(conditions)

            sql += " ORDER BY date"

            c.execute(sql, params)
            rows = c.fetchall()

            if not rows:
                self.tip_label.text = "没有符合条件的记录，无法导出"
                return

            # 写入Excel
            wb = Workbook()
            ws = wb.active
            ws.title = '导出结果'
            headers = ['时间', '收支', '金额', '事件', '相关方', '交易平台', '其他']
            ws.append(headers)

            for row in rows:
                # row 顺序与 SELECT 一致
                ws.append(row)

            wb.save(full_path)
            self.tip_label.text = f"导出成功！\n文件保存至：{full_path}"
            self.tip_label.color = (0, 0.7, 0, 1)
        except Exception as e:
            self.tip_label.text = f"导出失败：{str(e)}"
            self.tip_label.color = (1, 0, 0, 1)
        finally:
            conn.close()

    def go_back(self, instance):
        self.manager.current = 'main'


# ---------- 应用入口 ----------
class MyApp(App):
    def build(self):
        self.bills_dir = os.path.join(os.path.dirname(__file__), 'bills')
        os.makedirs(self.bills_dir, exist_ok=True)

        self.bill_files = []          # 存储 .db 文件名列表
        self.current_bill = None
        self.refresh_bill_list()

        sm = ScreenManager()
        sm.add_widget(MainScreen(name='main'))
        sm.add_widget(ViewEventScreen(name='view_event'))
        sm.add_widget(RecordScreen(name='record'))
        sm.add_widget(ImportScreen(name='import'))
        sm.add_widget(ExportScreen(name='export'))
        return sm

    def refresh_bill_list(self):
        self.bill_files = [f for f in os.listdir(self.bills_dir) if f.endswith('.db')]
        if not self.bill_files:
            default_name = '我的账单.db'
            default_path = os.path.join(self.bills_dir, default_name)
            init_database(default_path)
            self.bill_files = [default_name]
        if self.current_bill is None or self.current_bill not in self.bill_files:
            self.current_bill = self.bill_files[0]

    def create_new_bill(self, name):
        if not name.strip():
            return False, "文件名不能为空"
        filename = name.strip() + '.db'
        filepath = os.path.join(self.bills_dir, filename)
        if os.path.exists(filepath):
            return False, "文件已存在"
        init_database(filepath)
        self.refresh_bill_list()
        self.current_bill = filename
        return True, "创建成功"

    def rename_current_bill(self, new_name):
        if not self.current_bill:
            return False, "没有选中任何账单"
        new_name = new_name.strip()
        if not new_name:
            return False, "名称不能为空"
        if new_name.lower().endswith('.db'):
            new_name = new_name[:-3]
        new_filename = new_name + '.db'
        if new_filename == self.current_bill:
            return False, "新名称与当前名称相同"
        new_path = os.path.join(self.bills_dir, new_filename)
        if os.path.exists(new_path):
            return False, "该名称已存在，请换一个"
        old_path = os.path.join(self.bills_dir, self.current_bill)
        try:
            os.rename(old_path, new_path)
        except Exception as e:
            return False, f"重命名失败：{str(e)}"
        self.current_bill = new_filename
        self.refresh_bill_list()
        return True, "重命名成功"

    def delete_current_bill(self):
        if not self.current_bill:
            return False, "没有选中任何账单"
        if len(self.bill_files) == 1:
            return False, "至少需要保留一个账单，无法删除最后一个"
        filepath = os.path.join(self.bills_dir, self.current_bill)
        try:
            os.remove(filepath)
        except Exception as e:
            return False, f"删除失败：{str(e)}"
        self.refresh_bill_list()
        self.current_bill = self.bill_files[0]
        return True, "删除成功"

    def get_current_db_path(self):
        return os.path.join(self.bills_dir, self.current_bill)


if __name__ == '__main__':

    MyApp().run()
